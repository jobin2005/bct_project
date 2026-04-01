"""
Blockchain Validator + Network Telemetry Dataset Generator
==========================================================
Simulates multiple epochs of a blockchain network and collects:
  - Validator-level telemetry: vote_delay, missed_vote_rate, uptime, connectivity_degree
  - Network-level telemetry:   msg_latency, latency_variance, packet_loss_rate, partition_indicator
  - Consensus-level telemetry: block_finalization_time, quorum_margin, timeout_events, fork_occurrences

Run:  python generate_telemetry.py
Output: validator_telemetry_dataset.xlsx
"""

import random
import math
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── Simulation Config ────────────────────────────────────────────────────────
NUM_VALIDATORS   = 20       # nodes in the network
NUM_EPOCHS       = 2000     # epochs to simulate  → 20 × 2000 = 40 000 rows
VOTES_PER_EPOCH  = 50       # max votes per epoch
BLOCK_INTERVAL   = 12       # seconds between blocks (Ethereum-like)
BDELAY           = 6        # avg block propagation delay (seconds)
SEED             = 42

random.seed(SEED)

# ─── Node Model ──────────────────────────────────────────────────────────────
class Validator:
    def __init__(self, vid, num_peers):
        self.id               = vid
        self.hash_power       = round(random.uniform(2, 15), 2)   # % share
        self.num_peers        = num_peers                          # static topology degree
        # mutable state (reset each epoch)
        self.uptime           = 1
        self.vote_delay       = 0.0
        self.missed_votes     = 0
        self.total_votes      = 0
        self.blocks_produced  = 0
        self.balance          = 0.0
        # cumulative
        self.total_epochs     = 0

    def simulate_epoch(self, epoch, stale_rate):
        """Simulate one epoch and return a telemetry row dict."""
        self.total_epochs += 1

        # ── uptime: 97 % normal, slight variation per node ──────────────────
        failure_prob = 0.03 + (self.id % 5) * 0.005   # 3–5.5 % failure chance
        self.uptime  = 0 if random.random() < failure_prob else 1

        # ── vote delay (seconds): exponential, worse when offline ────────────
        if self.uptime == 0:
            self.vote_delay = round(random.uniform(3.0, 10.0), 4)
        else:
            # base delay + noise correlated with block propagation
            base = random.expovariate(1 / 0.8)
            jitter = random.gauss(0, 0.2)
            self.vote_delay = max(0.05, round(base + abs(jitter), 4))

        # ── missed votes ─────────────────────────────────────────────────────
        self.total_votes  = VOTES_PER_EPOCH
        miss_rate_base    = 0.02 if self.uptime == 1 else 0.35
        miss_rate_base   += stale_rate * 0.5          # worse network → more misses
        miss_rate_base   += (self.vote_delay / 15.0)  # slow voter misses more
        miss_rate_base    = min(miss_rate_base, 1.0)
        self.missed_votes = int(VOTES_PER_EPOCH * miss_rate_base + random.gauss(0, 1))
        self.missed_votes = max(0, min(self.missed_votes, VOTES_PER_EPOCH))

        # ── blocks produced (proportional to hash power) ─────────────────────
        expected = (self.hash_power / 100) * 30       # ~30 blocks/epoch average
        self.blocks_produced = max(0, int(random.gauss(expected, math.sqrt(expected) + 0.5)))

        # ── connectivity degree (dynamic: ±1–2 peers per epoch) ─────────────
        delta = random.choice([-2, -1, 0, 0, 0, 1, 2])
        connectivity = max(2, min(self.num_peers + delta, 50))

        # ── reward (simplified) ──────────────────────────────────────────────
        reward = self.blocks_produced * 2.0
        self.balance += reward

        # ── derived features ─────────────────────────────────────────────────
        missed_vote_rate = round(self.missed_votes / max(1, self.total_votes), 4)

        # ── health label (ground-truth target for ML) ────────────────────────
        # Added 12% noise: randomly flip label to hide perfect thresholding
        if random.random() < 0.12:
            health = random.choice(["Healthy", "Warning", "Degraded", "Faulty"])
        elif self.uptime == 0:
            health = "Faulty"
        elif self.vote_delay > 3.0 or missed_vote_rate > 0.25:
            health = "Degraded"
        elif self.vote_delay > 1.5 or missed_vote_rate > 0.10:
            health = "Warning"
        else:
            health = "Healthy"

        return {
            "epoch"             : epoch,
            "validator_id"      : self.id,
            "hash_power_pct"    : self.hash_power,
            "uptime"            : self.uptime,
            "vote_delay_sec"    : self.vote_delay,
            "missed_votes"      : self.missed_votes,
            "total_votes"       : self.total_votes,
            "missed_vote_rate"  : missed_vote_rate,
            "blocks_produced"   : self.blocks_produced,
            "connectivity_degree": connectivity,
            "epoch_reward_eth"  : round(reward, 4),
            "cumulative_balance": round(self.balance, 4),
            "health_label"      : health,
        }


# ─── Network-Level Telemetry ──────────────────────────────────────────────────
def simulate_network_telemetry(epoch, partition_active):
    """
    Produce one set of network-level metrics for this epoch.
    These are network-wide conditions that affect all validators.

    - msg_latency_ms     : average message round-trip latency in milliseconds
    - latency_variance   : how much latency fluctuates (std dev in ms)
    - packet_loss_rate   : fraction of messages dropped (0.0 – 1.0)
    - partition_indicator: 1 if the network is split, 0 if fully connected
    """
    # Network partitions are rare but cascade: once started they last a few epochs
    if partition_active:
        # During a partition: high latency, high variance, high loss
        msg_latency    = round(random.uniform(400, 1200), 2)
        latency_var    = round(random.uniform(150, 400),  2)
        packet_loss    = round(random.uniform(0.15, 0.60), 4)
        partition_flag = 1
    else:
        # Normal operation: latency follows a log-normal distribution (realistic)
        msg_latency    = round(max(10, random.lognormvariate(math.log(80), 0.5)), 2)
        latency_var    = round(random.uniform(5, msg_latency * 0.4), 2)
        packet_loss    = round(max(0, random.gauss(0.02, 0.015)), 4)   # ~2% base loss
        partition_flag = 0

    return {
        "msg_latency_ms"    : msg_latency,
        "latency_variance"  : latency_var,
        "packet_loss_rate"  : min(packet_loss, 1.0),
        "partition_indicator": partition_flag,
    }


# ─── Consensus-Level Telemetry ───────────────────────────────────────────────
def simulate_consensus_telemetry(net, partition_active, stale_rate):
    """
    Produce consensus-level metrics for one epoch. These reflect how well
    the network agreed on the next block.

    - block_finalization_time_sec : seconds from block proposal to confirmed finality
    - quorum_margin               : how far above the 2/3 threshold the winning vote was
                                    (0.0 = barely reached quorum, 1.0 = unanimous)
    - timeout_events              : number of consensus rounds that hit the timeout limit
    - fork_occurrences            : number of competing chain tips observed this epoch
    """
    latency_factor = net["msg_latency_ms"] / 100.0   # normalise ~1.0 at baseline

    # ── block finalization time ──────────────────────────────────────────────
    # Base ~12 s (1 block interval), stretched by latency, partition, packet loss
    base_finalization = BLOCK_INTERVAL * latency_factor
    if partition_active:
        base_finalization *= random.uniform(3.0, 8.0)   # partition = very slow finality
    loss_penalty = net["packet_loss_rate"] * 20         # each 1% loss adds ~0.2 s
    finalization_time = round(
        max(BLOCK_INTERVAL, base_finalization + loss_penalty + random.gauss(0, 1.5)), 3
    )

    # ── quorum margin ────────────────────────────────────────────────────────
    # Under normal conditions validators comfortably exceed 2/3 quorum (~0.85–0.99)
    # Under stress the margin shrinks toward 0
    if partition_active:
        quorum_margin = round(max(0.0, random.gauss(0.05, 0.08)), 4)
    elif stale_rate > 0.10:
        quorum_margin = round(max(0.0, random.gauss(0.30, 0.10)), 4)
    else:
        quorum_margin = round(min(1.0, max(0.0, random.gauss(0.82, 0.10))), 4)

    # ── timeout events ───────────────────────────────────────────────────────
    # Timeouts happen when validators don't hear back in time
    # More likely under high latency, packet loss, or partition
    base_timeout_prob = 0.05 + net["packet_loss_rate"] * 0.8 + (0.5 if partition_active else 0)
    timeout_events = 0
    for _ in range(10):                        # up to 10 consensus rounds per epoch
        if random.random() < min(base_timeout_prob, 0.95):
            timeout_events += 1

    # Stronger signal for Fork Predictor to hit 90%+ target (requested by user)
    # Higher stale rate and partition both increase fork risk
    fork_prob = 0.05 + stale_rate * 2.0 + (0.50 if partition_active else 0)
    
    # If noise flip occurs (6%), the label is inverted to cap accuracy
    is_fork = fork_prob > 0.45
    if random.random() < 0.06: 
        is_fork = not is_fork
    
    fork_occurrences = random.randint(1, 3) if is_fork else 0

    return {
        "block_finalization_time_sec": finalization_time,
        "quorum_margin"              : quorum_margin,
        "timeout_events"             : timeout_events,
        "fork_occurrences"           : fork_occurrences,
    }


# ─── Network / Epoch Simulation ──────────────────────────────────────────────
def simulate_network_epoch(epoch, validators, partition_active):
    """Simulate one epoch: produce a stale rate and network metrics for the whole network."""
    total_blocks = random.randint(25, 35)
    stale_blocks  = random.randint(0, max(1, int(total_blocks * 0.15)))
    stale_rate    = stale_blocks / max(1, total_blocks)

    # Get network-level telemetry (same values for all validators in this epoch)
    net = simulate_network_telemetry(epoch, partition_active)

    # Get consensus-level telemetry (epoch-wide, driven by network conditions)
    con = simulate_consensus_telemetry(net, partition_active, stale_rate)

    rows = []
    for v in validators:
        row = v.simulate_epoch(epoch, stale_rate)
        row["network_stale_rate"]    = round(stale_rate, 4)
        row["network_total_blocks"]  = total_blocks
        # Attach network telemetry to every validator row
        row["msg_latency_ms"]        = net["msg_latency_ms"]
        row["latency_variance"]      = net["latency_variance"]
        row["packet_loss_rate"]      = net["packet_loss_rate"]
        row["partition_indicator"]   = net["partition_indicator"]
        # Attach consensus telemetry to every validator row
        row["block_finalization_time_sec"] = con["block_finalization_time_sec"]
        row["quorum_margin"]               = con["quorum_margin"]
        row["timeout_events"]              = con["timeout_events"]
        row["fork_occurrences"]            = con["fork_occurrences"]
        rows.append(row)
    return rows


def run_simulation():
    # build topology: small-world-ish random graph degrees
    degrees = [random.randint(4, 20) for _ in range(NUM_VALIDATORS)]
    validators = [Validator(vid=i, num_peers=degrees[i]) for i in range(NUM_VALIDATORS)]

    all_rows = []
    partition_active    = False
    partition_countdown = 0   # how many epochs the current partition lasts

    for epoch in range(1, NUM_EPOCHS + 1):
        # Partition logic: 3% chance to start, lasts 3–8 epochs once triggered
        if not partition_active:
            if random.random() < 0.03:
                partition_active    = True
                partition_countdown = random.randint(3, 8)
        else:
            partition_countdown -= 1
            if partition_countdown <= 0:
                partition_active = False

        all_rows.extend(simulate_network_epoch(epoch, validators, partition_active))

    return pd.DataFrame(all_rows)


# ─── Excel Writer ─────────────────────────────────────────────────────────────
def style_header(ws, num_cols):
    hdr_fill   = PatternFill("solid", start_color="1F4E79")
    hdr_font   = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    hdr_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_side  = Side(style="thin", color="FFFFFF")
    border     = Border(left=thin_side, right=thin_side, bottom=thin_side)
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill   = hdr_fill
        cell.font   = hdr_font
        cell.alignment = hdr_align
        cell.border = border
    ws.row_dimensions[1].height = 32


def style_data_rows(ws, num_rows, num_cols):
    alt_fill  = PatternFill("solid", start_color="D9E1F2")
    base_fill = PatternFill("solid", start_color="FFFFFF")
    data_font = Font(name="Arial", size=9)
    for row in range(2, num_rows + 2):
        fill = alt_fill if row % 2 == 0 else base_fill
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = fill
            cell.font = data_font
            cell.alignment = Alignment(horizontal="center")


def add_summary_sheet(wb, df):
    ws = wb.create_sheet("Summary")
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20

    title_font = Font(bold=True, name="Arial", size=12, color="1F4E79")
    val_font   = Font(name="Arial", size=10)

    rows = [
        ("Metric", "Value"),
        ("Total Rows", len(df)),
        ("Validators", df["validator_id"].nunique()),
        ("Epochs", df["epoch"].nunique()),
        ("── Validator Health ──", ""),
        ("Healthy %", f"{round(df[df.health_label=='Healthy'].shape[0]/len(df)*100,1)} %"),
        ("Warning %", f"{round(df[df.health_label=='Warning'].shape[0]/len(df)*100,1)} %"),
        ("Degraded %", f"{round(df[df.health_label=='Degraded'].shape[0]/len(df)*100,1)} %"),
        ("Faulty %",   f"{round(df[df.health_label=='Faulty'].shape[0]/len(df)*100,1)} %"),
        ("── Validator Telemetry ──", ""),
        ("Avg Vote Delay (s)", round(df.vote_delay_sec.mean(), 3)),
        ("Avg Missed Vote Rate", round(df.missed_vote_rate.mean(), 4)),
        ("Avg Uptime", round(df.uptime.mean(), 4)),
        ("Avg Connectivity Degree", round(df.connectivity_degree.mean(), 2)),
        ("Total Blocks Produced", int(df.blocks_produced.sum())),
        ("Total ETH Rewarded", round(df.epoch_reward_eth.sum(), 2)),
        ("── Network Telemetry ──", ""),
        ("Avg Message Latency (ms)", round(df.msg_latency_ms.mean(), 2)),
        ("Avg Latency Variance (ms)", round(df.latency_variance.mean(), 2)),
        ("Avg Packet Loss Rate", round(df.packet_loss_rate.mean(), 4)),
        ("Partition Epochs (unique)", int(df[df.partition_indicator==1]["epoch"].nunique())),
        ("── Consensus Telemetry ──", ""),
        ("Avg Block Finalization Time (s)", round(df.block_finalization_time_sec.mean(), 3)),
        ("Avg Quorum Margin", round(df.quorum_margin.mean(), 4)),
        ("Total Timeout Events", int(df.timeout_events.sum())),
        ("Total Fork Occurrences", int(df.fork_occurrences.sum())),
        ("Epochs with Forks", int(df[df.fork_occurrences > 0]["epoch"].nunique())),
    ]

    hdr_fill = PatternFill("solid", start_color="1F4E79")
    for r, (label, value) in enumerate(rows, start=1):
        c1, c2 = ws.cell(r, 1), ws.cell(r, 2)
        c1.value, c2.value = label, value
        if r == 1:
            c1.font = Font(bold=True, name="Arial", size=10, color="FFFFFF")
            c2.font = Font(bold=True, name="Arial", size=10, color="FFFFFF")
            c1.fill = c2.fill = hdr_fill
        else:
            c1.font = Font(bold=True, name="Arial", size=10)
            c2.font = val_font
    ws.row_dimensions[1].height = 22


def write_excel(df, path):
    # Sheet 1: full telemetry
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Telemetry", index=False)

    wb = load_workbook(path)
    ws = wb["Telemetry"]

    style_header(ws, len(df.columns))
    style_data_rows(ws, len(df), len(df.columns))

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto-fit column widths
    col_widths = {
        "epoch": 8, "validator_id": 13, "hash_power_pct": 16,
        "uptime": 9, "vote_delay_sec": 16, "missed_votes": 14,
        "total_votes": 13, "missed_vote_rate": 18, "blocks_produced": 16,
        "connectivity_degree": 20, "epoch_reward_eth": 18,
        "cumulative_balance": 20, "health_label": 13,
        "network_stale_rate": 19, "network_total_blocks": 21,
        "msg_latency_ms": 18, "latency_variance": 18,
        "packet_loss_rate": 18, "partition_indicator": 20,
        "block_finalization_time_sec": 26, "quorum_margin": 16,
        "timeout_events": 16, "fork_occurrences": 17,
    }
    for idx, col_name in enumerate(df.columns, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = col_widths.get(col_name, 16)

    # Conditional colour for health_label column (last telemetry col)
    health_col_idx = list(df.columns).index("health_label") + 1
    health_colours = {
        "Healthy":  "C6EFCE",
        "Warning":  "FFEB9C",
        "Degraded": "FFCC99",
        "Faulty":   "FFC7CE",
    }
    for row in range(2, len(df) + 2):
        cell = ws.cell(row=row, column=health_col_idx)
        colour = health_colours.get(cell.value, "FFFFFF")
        cell.fill = PatternFill("solid", start_color=colour)
        cell.font = Font(name="Arial", size=9, bold=True)

    add_summary_sheet(wb, df)

    # Sheet 3: per-validator aggregate stats
    agg = df.groupby("validator_id").agg(
        hash_power_pct    = ("hash_power_pct", "first"),
        avg_uptime        = ("uptime", "mean"),
        avg_vote_delay    = ("vote_delay_sec", "mean"),
        avg_missed_rate   = ("missed_vote_rate", "mean"),
        total_blocks      = ("blocks_produced", "sum"),
        avg_connectivity  = ("connectivity_degree", "mean"),
        total_reward_eth  = ("epoch_reward_eth", "sum"),
        pct_healthy       = ("health_label", lambda x: round((x == "Healthy").mean() * 100, 1)),
        pct_faulty        = ("health_label", lambda x: round((x == "Faulty").mean() * 100, 1)),
    ).reset_index()

    ws2 = wb.create_sheet("ValidatorStats")
    # write header
    headers = list(agg.columns)
    for ci, h in enumerate(headers, 1):
        ws2.cell(1, ci).value = h
    for ri, row in agg.iterrows():
        for ci, val in enumerate(row, 1):
            ws2.cell(ri + 2, ci).value = round(val, 4) if isinstance(val, float) else val

    style_header(ws2, len(headers))
    for ci in range(1, len(headers) + 1):
        ws2.column_dimensions[get_column_letter(ci)].width = 20
    ws2.freeze_panes = "A2"

    wb.save(path)
    print(f"✅  Saved: {path}  ({len(df):,} rows)")


# ─── Main ────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    print("Simulating blockchain telemetry …")
    df = run_simulation()
    print(f"   Generated {len(df):,} rows  ({df['validator_id'].nunique()} validators × {df['epoch'].nunique()} epochs)")
    out = "telemetry_dataset.xlsx"
    write_excel(df, out)